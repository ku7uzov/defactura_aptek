import time
import csv
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
import sys
import tempfile
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
# from webdriver_manager.chrome import ChromeDriverManager

def create_driver() -> webdriver.Chrome:
    options = Options()
    # options.add_argument("--headless")  # включи при необходимости
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    user_data_dir = tempfile.mkdtemp()
    options.add_argument(f"--user-data-dir={user_data_dir}")
    driver =  webdriver.Chrome(options=options)
    # driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    # временная уникальная папка для профиля

    return driver


# Установить отображение 100 аптек на странице
def set_items_per_page(driver: webdriver.Chrome) -> None:
    try:
        try:
            bottom_notice = driver.find_element(By.CLASS_NAME, "bottom-notice-close")
            driver.execute_script("arguments[0].click();", bottom_notice)
            print(" Баннер закрыт")
            time.sleep(1)
        except NoSuchElementException:
            pass

        select_buttons = driver.find_elements(By.ID, "paging")
        if not select_buttons:
            print(" Элемент #paging не найден — возможно, показ всех аптек по умолчанию.")
            return

        select_button = select_buttons[0]
        driver.execute_script("arguments[0].scrollIntoView(true);", select_button)
        driver.execute_script("arguments[0].click();", select_button)

        option_100 = driver.find_element(By.ID, "tw3")
        driver.execute_script("arguments[0].click();", option_100)
        # time.sleep(1)
    except Exception as e:
        print(" Ошибка при выборе количества аптек:", e)


# Сбор информации об аптеках на странице
from selenium.common.exceptions import StaleElementReferenceException


def get_pharmacy_info(driver: webdriver.Chrome) -> list[list[str]]:
    pharmacies = []
    rows = driver.find_elements(By.CSS_SELECTOR, "tr.tr-border")

    for index in range(len(rows)):
        try:
            # Получаем строку снова, чтобы избежать stale reference
            row = driver.find_elements(By.CSS_SELECTOR, "tr.tr-border")[index]

            name = row.find_element(By.CSS_SELECTOR, ".pharm-name a").text.strip()
            address = row.find_element(By.CSS_SELECTOR, ".address .text-wrap span").text.strip()
            phone = row.find_element(By.CSS_SELECTOR, ".phone .text-wrap a").text.strip()
            price = row.find_element(By.CSS_SELECTOR, ".price .price-value").text.strip()
            pharmacies.append([name, address, phone, price])
        except (NoSuchElementException, StaleElementReferenceException):
            continue

    return pharmacies


# Переход на следующую страницу
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

def click_next(driver: webdriver.Chrome) -> bool:
    try:
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".table-pagination-next"))
        )

        next_btns = driver.find_elements(By.CSS_SELECTOR, ".table-pagination-next a")
        if not next_btns:
            print(" Кнопка 'Вперёд' не найдена. Страницы закончились.")
            return False

        next_btn = next_btns[0]
        if not next_btn.is_displayed():
            print(" Кнопка 'Вперёд' не видна. Страницы закончились.")
            return False

        driver.execute_script("arguments[0].scrollIntoView(true);", next_btn)
        driver.execute_script("window.scrollBy(0, 100);")
        driver.execute_script("arguments[0].click();", next_btn)
        return True
    except Exception as e:
        print(" Ошибка при переходе на следующую страницу:", e)
        return False



# Чтение всех аптек
def read_pharmacies(file_name: str) -> set[tuple[str, str]]:
    with open(file_name, newline='', encoding='utf-8') as f:
        reader = csv.reader(f)
        next(reader, None)
        return {(row[0].strip(), row[1].strip()) for row in reader if row}


# Чтение аптек, у которых есть препарат
# Чтение второго файла (аптеки с препаратом)
def read_pharmacies_with_drug(file_name):
    with open(file_name, newline='', encoding='utf-8') as f:
        reader = csv.reader(f)
        next(reader, None)  # ⬅️ пропустить заголовок
        pharmacies = {(row[0].strip(), row[1].strip()) for row in reader if row}
    return pharmacies


# Основной алгоритм
# Основной алгоритм
def compare_pharmacies():
    # Чтение всех аптек
    all_pharmacies = read_pharmacies('all_pharmacies.csv')

    # Чтение аптек с препаратом
    pharmacies_with_drug = read_pharmacies_with_drug('pharmacies_with_drug.csv')

    # 🔍 Отладочная информация
    print(f" Всего строк в all_pharmacies.csv (включая заголовок): {sum(1 for _ in open('all_pharmacies.csv', encoding='utf-8'))}")
    print(f" Уникальных аптек (по названию и адресу): {len(all_pharmacies)}")
    print(f" Аптек, в которых есть препарат: {len(pharmacies_with_drug)}")

    # Находим аптеки, в которых нет препарата
    pharmacies_without_drug = all_pharmacies - pharmacies_with_drug

    print(f" Всего аптек без препарата: {len(pharmacies_without_drug)}")

    # Чтение дополнительной информации о всех аптеках (из первого файла)
    pharmacies_info = {}
    with open('all_pharmacies.csv', newline='', encoding='utf-8') as f:
        reader = csv.reader(f)
        next(reader, None)  # Пропустить заголовок
        for row in reader:
            name = row[0].strip()
            address = row[1].strip()
            work_time = row[2].strip()
            phone = row[3].strip()
            pharmacies_info[(name, address)] = {
                "work_time": work_time,
                "phone": phone,
            }

    # Записываем результат в новый CSV
    output_data = []
    with open('pharmacies_without_drug.csv', "w", newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(["Название аптеки", "Адрес", "Телефон", "Время работы", "Цена"])

        for pharmacy in pharmacies_without_drug:
            if pharmacy in pharmacies_info:
                info = pharmacies_info[pharmacy]
                row = [pharmacy[0], pharmacy[1], info["phone"], info["work_time"], ""]
                writer.writerow(row)
                output_data.append(row)

    print(" Аптеки, в которых нет препарата, сохранены в pharmacies_without_drug.csv")

    #  Добавляем сохранение в Excel
    save_to_excel("pharmacies_without_drug.xlsx", output_data, ["Название аптеки", "Адрес", "Телефон", "Время работы"])

    #  Сохраняем результат в БД



from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def save_to_excel(filename, data, headers):
    wb = Workbook()
    ws = wb.active
    ws.title = "Аптеки"

    # Заголовки
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = max(len(header) + 2, 15)

    # Данные
    for row_num, row_data in enumerate(data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    wb.save(filename)
    print(f"Данные сохранены в {filename}")


# Главная функция
def parser(item_id):
    driver = create_driver()
    BASE_URL = f"https://tabletka.by/result/?ls={item_id}"

    # print(" Открываем страницу...")
    driver.get(BASE_URL)
    time.sleep(2)
    print('sleep 2 seconds')

    # print("️ Устанавливаем показ по 100...")
    set_items_per_page(driver)

    result = []
    page = 1

    while True:
        print(f" Парсим страницу {page}...")
        result.extend(get_pharmacy_info(driver))
        if not click_next(driver):
            break
        time.sleep(1)
        page += 1

    print(f" Собрано аптек: {len(result)}")

    with open("pharmacies_with_drug.csv", "w", newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(["Название аптеки", "Адрес", "Телефон", "Цена"])
        writer.writerows(result)

    compare_pharmacies()
    print(" Данные сохранены в pharmacies_with_drug.csv")
    driver.quit()




# if __name__ == "__main__":
#     # Получаем item_id из командной строки
#     if len(sys.argv) != 2:
#         print(" Ошибка: требуется передать item_id как аргумент командной строки.")
#         sys.exit(1)
#
#     try:
#         item_id = int(sys.argv[1])  # Преобразуем в int (если это число)
#     except ValueError:
#         print(" Ошибка: item_id должен быть числом.")
#         sys.exit(1)
#
#     # Запускаем парсер с переданным item_id
#     parser(item_id)



