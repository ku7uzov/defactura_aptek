import requests
from bs4 import BeautifulSoup
import re





def parse_tablets(name, region):
    url = f"https://tabletka.by/search/?request={name}&region={region}"
    response = requests.get(url)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, 'html.parser')


    warnings = [notice.text.strip() for notice in soup.select("div.notice.notice--warning")]
    dangers = [notice.text.strip() for notice in soup.select("div.notice.notice--danger")]

    # Фильтруем предупреждения, удаляя "По заданному фильтру не найдено"
    warnings = [w for w in warnings if w != "Данных по заданному фильтру не найдено."]

    # Проверяем, есть ли сообщение о ненайденных совпадениях
    no_results_message = None
    suggestions = []
    no_results = soup.select_one("h2.title-h2.page-title")
    if no_results and "По вашему запросу совпадений не найдено" in no_results.text:
        no_results_message = "По вашему запросу совпадений не найдено. Возможно, вы искали:"
        suggestions = [s.text.strip() for s in soup.select("div.link-block-wrap a .bttn.link-block")]

    # Извлекаем данные о таблетках
    tablets = []
    for row in soup.select("tbody.tbody-base-tbl tr.tr-border"):
        item_id_tag = row.select_one("td.btn .heart-icon")
        item_id = item_id_tag["itemid"] if item_id_tag else None  # Извлекаем itemID


        name_tag = row.select_one("td.name .tooltip-info-header a")
        form_tag = row.select_one("td.form .tooltip-info-header a")
        produce = row.select_one("td.produce .tooltip-info-header a")
        price_tag = row.select_one("td.price .price-value")
        mnn_tag = row.select_one("td.name .capture a")  # Парсим ссылку на МНН
        recipe_tag = row.select_one("td.form .capture")  # Добавим для получения информации о рецепте

        # Извлекаем данные о количестве аптек и ссылке
        pharmacy_info_tag = row.select_one("td.price .capture a")
        if pharmacy_info_tag:
            # Ищем первое число в тексте, например: "в 168 аптеках"
            match = re.search(r"\d+", pharmacy_info_tag.text)
            pharmacy_info = int(match.group()) if match else 0
        else:
            pharmacy_info = None

        # pharmacy_info = pharmacy_info_tag.text.strip() if pharmacy_info_tag else None
        pharmacy_link = pharmacy_info_tag['href'] if pharmacy_info_tag else None

        if name_tag and form_tag and price_tag and produce:
            item_id_tag = row.select_one("td.btn .heart-icon")
            item_id = item_id_tag["itemid"] if item_id_tag else None  # Извлекаем itemID

            name = name_tag.text.strip()
            form = form_tag.text.strip()
            produce = produce.text.strip()
            price_text = price_tag.text.strip()
            form_link = form_tag["href"].replace("®", "&reg")  # Берем ссылку на форму
            mnn_link = ""
            mnn_name = ""
            try:
                mnn_link = mnn_tag['href'].replace("®", "&reg")
                mnn_name = mnn_tag.text.strip() if mnn_tag else None
            except:
                pass

            recipe = recipe_tag.text.strip() if recipe_tag else "Не указано"

            # Извлекаем диапазон цен (например, 17.56 ... 18.42)
            price_match = re.search(r"([\d.]+)\s*\.\.\.\s*([\d.]+)\s*р\.", price_text)

            if price_match:
                price_from = float(price_match.group(1))
                price_to = float(price_match.group(2))
            else:
                # Если диапазона нет, пытаемся найти одну цену
                single_price_match = re.search(r"([\d.]+)\s*р\.", price_text)
                price_from = float(single_price_match.group(1)) if single_price_match else None
                price_to = None  # Указываем, что второй цены нет

            tablets.append({
                "item_id" :item_id,
                "name": name,
                "form": form,
                "produce": produce,
                "form_link": form_link,
                "price_from": price_from,
                "price_to": price_to,
                "mnn_link": mnn_link,
                "mnn_name": mnn_name,
                "recipe": recipe,
                "pharmacy_diff" : 4283 - int(pharmacy_info),
                "pharmacy_info": pharmacy_info,  # Добавляем количество аптек
                "pharmacy_link": pharmacy_link   # Добавляем ссылку на результаты
            })


    # Сортируем список по минимальной цене
    tablets.sort(key=lambda x: x["price_from"] if x["price_from"] is not None else float('inf'))


    return {
        "tablets": tablets,
        "warnings": warnings,
        "dangers": dangers,
        "no_results_message": no_results_message,
        "suggestions": suggestions
    }




import time
import csv
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
import sys
import tempfile
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
# from webdriver_manager.chrome import ChromeDriverManager
import os

def create_driver() -> webdriver.Chrome:
    options = Options()
    # options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')

    # Использование уникального имени для папки
    # user_data_dir = f"/tmp/chrome_profile_{int(time.time())}"
    # os.makedirs(user_data_dir, exist_ok=True)
    #
    # options.add_argument(f"--user-data-dir={user_data_dir}")

    # Запуск драйвера с указанными опциями
    # driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver = webdriver.Chrome(options=options)

    return driver


# Установить отображение 100 аптек на странице
def set_items_per_page(driver: webdriver.Chrome) -> None:
    try:
        try:
            accept_button = driver.find_element(By.CLASS_NAME, "allow-btn")
            driver.execute_script("arguments[0].click();", accept_button)
            print("Кнопка 'Принимаю' нажата")
            time.sleep(1)
        except NoSuchElementException:
            print("Кнопка 'Принимаю' не найдена")

        try:
            bottom_notice = driver.find_element(By.CLASS_NAME, "bottom-notice-close")
            driver.execute_script("arguments[0].click();", bottom_notice)
            print("Баннер закрыт")
            time.sleep(1)
        except NoSuchElementException:
            pass

        select_buttons = driver.find_elements(By.ID, "paging")
        if not select_buttons:
            print("Элемент #paging не найден — возможно, показ всех аптек по умолчанию.")
            return

        select_button = select_buttons[0]
        driver.execute_script("arguments[0].scrollIntoView(true);", select_button)
        driver.execute_script("arguments[0].click();", select_button)

        option_100 = driver.find_element(By.ID, "tw3")
        driver.execute_script("arguments[0].click();", option_100)
        # time.sleep(1)
    except Exception as e:
        print(" Ошибка при выборе количества аптек:", e)


# Сбор информации об аптеках на странице
from selenium.common.exceptions import StaleElementReferenceException


def get_pharmacy_info(driver: webdriver.Chrome) -> list[list[str]]:
    pharmacies = []
    rows = driver.find_elements(By.CSS_SELECTOR, "tr.tr-border")

    for index in range(len(rows)):
        try:
            # Получаем строку снова, чтобы избежать stale reference
            row = driver.find_elements(By.CSS_SELECTOR, "tr.tr-border")[index]

            name = row.find_element(By.CSS_SELECTOR, ".pharm-name a").text.strip()
            address = row.find_element(By.CSS_SELECTOR, ".address .text-wrap span").text.strip()
            phone = row.find_element(By.CSS_SELECTOR, ".phone .text-wrap a").text.strip()
            price = row.find_element(By.CSS_SELECTOR, ".price .price-value").text.strip()
            pharmacies.append([name, address, phone, price])
        except (NoSuchElementException, StaleElementReferenceException):
            continue

    return pharmacies


# Переход на следующую страницу
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

def click_next(driver: webdriver.Chrome) -> bool:
    try:

        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".table-pagination-next"))
        )

        next_btns = driver.find_elements(By.CSS_SELECTOR, ".table-pagination-next a")
        if not next_btns:
            print(" Кнопка 'Вперёд' не найдена. Страницы закончились.")
            return False

        next_btn = next_btns[0]
        if not next_btn.is_displayed():
            print(" Кнопка 'Вперёд' не видна. Страницы закончились.")
            return False

        driver.execute_script("arguments[0].scrollIntoView(true);", next_btn)
        driver.execute_script("window.scrollBy(0, 100);")
        driver.execute_script("arguments[0].click();", next_btn)
        return True
    except Exception as e:
        print(" Ошибка при переходе на следующую страницу:", e)
        return False



# Чтение всех аптек
def read_pharmacies(file_name: str) -> set[tuple[str, str]]:
    with open(file_name, newline='', encoding='utf-8') as f:
        reader = csv.reader(f)
        next(reader, None)
        return {(row[0].strip(), row[1].strip()) for row in reader if row}


# Чтение аптек, у которых есть препарат
# Чтение второго файла (аптеки с препаратом)
def read_pharmacies_with_drug(file_name):
    with open(file_name, newline='', encoding='utf-8') as f:
        reader = csv.reader(f)
        next(reader, None)  # ⬅️ пропустить заголовок
        pharmacies = {(row[0].strip(), row[1].strip()) for row in reader if row}
    return pharmacies


# Основной алгоритм
# Основной алгоритм
def compare_pharmacies():
    # Чтение всех аптек
    # Чтение всех аптек
    all_pharmacies = read_pharmacies('all_pharmacies.csv')

    # Чтение аптек с препаратом
    pharmacies_with_drug = read_pharmacies_with_drug('pharmacies_with_drug.csv')

    # Отладочная информация
    print(f"Всего строк в all_pharmacies.csv: {sum(1 for _ in open('all_pharmacies.csv', encoding='utf-8'))}")
    print(f"Уникальных аптек: {len(all_pharmacies)}")
    print(f"Аптек с препаратом: {len(pharmacies_with_drug)}")

    # Находим аптеки, в которых нет препарата
    pharmacies_without_drug = all_pharmacies - pharmacies_with_drug

    print(f"Всего аптек без препарата: {len(pharmacies_without_drug)}")

    # Чтение информации о всех аптеках
    pharmacies_info = {}
    with open('all_pharmacies.csv', newline='', encoding='utf-8') as f:
        reader = csv.reader(f)
        next(reader, None)  # Пропустить заголовок
        for row in reader:
            if len(row) < 4:
                print(f"⚠️ Пропущена строка с недостаточным количеством элементов: {row}")
                continue
            name = row[0].strip()
            address = row[1].strip()
            work_time = row[2].strip()
            phone = row[3].strip()
            pharmacies_info[(name, address)] = {
                "work_time": work_time,
                "phone": phone,
            }

    # Запись результата в новый CSV
    output_data = []
    with open('pharmacies_without_drug.csv', "w", newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(["Название аптеки", "Адрес", "Телефон", "Время работы"])

        for pharmacy in pharmacies_without_drug:
            if pharmacy in pharmacies_info:
                info = pharmacies_info[pharmacy]
                row = [pharmacy[0], pharmacy[1], info["phone"], info["work_time"]]
                writer.writerow(row)
                output_data.append(row)

    print("Аптеки без препарата сохранены в pharmacies_without_drug.csv")

     # Добавляем сохранение в Excel
    save_to_excel("pharmacies_without_drug.xlsx", output_data, ["Название аптеки", "Адрес", "Телефон", "Время работы"])

    #  Сохраняем результат в БД



from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def save_to_excel(filename, data, headers):
    wb = Workbook()
    ws = wb.active
    ws.title = "Аптеки"

    # Заголовки
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = max(len(header) + 2, 15)

    # Данные
    for row_num, row_data in enumerate(data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    wb.save(filename)
    print(f"Данные сохранены в {filename}")


# Главная функция
def parser(item_id):

    driver = create_driver()
    try:
        BASE_URL = f"https://tabletka.by/result/?ls={item_id}"

        # print(" Открываем страницу...")
        driver.get(BASE_URL)
        time.sleep(2)
        print('sleep 2 seconds')

        # print("️ Устанавливаем показ по 100...")
        set_items_per_page(driver)

        result = []
        page = 1

        while True:
            print(f" Парсим страницу {page}...")
            result.extend(get_pharmacy_info(driver))
            if not click_next(driver):
                break
            time.sleep(1)
            page += 1

        print(f" Собрано аптек: {len(result)}")
        file_path = "pharmacies_with_drug.csv"

        with open(file_path, "w", newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(["Название аптеки", "Адрес", "Телефон"])
            writer.writerows(result)

        compare_pharmacies()
        print(" Данные сохранены в pharmacies_with_drug.csv")
    finally:
        driver.quit()


